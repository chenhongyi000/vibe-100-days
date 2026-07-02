import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.excel_cleaner import CleanOptions, clean_workbook, delete_duplicate_rows, is_blank, main


def make_dirty_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append([" name ", " city ", None, " amount "])
    ws.append([" Alice ", " Shanghai\nChina ", None, 100])
    ws.append([None, None, None, None])
    ws.append([" Bob ", " Beijing ", None, 200])
    ws.append([" Bob ", " Beijing ", None, 200])
    wb.save(path)


def test_clean_workbook_removes_blanks_and_normalizes_text(tmp_path: Path):
    source = tmp_path / "dirty.xlsx"
    output = tmp_path / "clean.xlsx"
    make_dirty_workbook(source)

    report = clean_workbook(CleanOptions(input_path=source, output_path=output, dedupe_rows=True))
    wb = load_workbook(output)
    ws = wb["Orders"]

    assert output.exists()
    assert ws.max_column == 3
    assert ws.max_row == 3
    assert ws["A1"].value == "name"
    assert ws["B2"].value == "Shanghai China"
    assert report.sheets[0].deleted_blank_rows == 1
    assert report.sheets[0].deleted_blank_columns == 1
    assert report.sheets[0].deleted_duplicate_rows == 1


def test_report_json_written(tmp_path: Path):
    source = tmp_path / "dirty.xlsx"
    output = tmp_path / "clean.xlsx"
    report_path = tmp_path / "report.json"
    make_dirty_workbook(source)

    clean_workbook(CleanOptions(input_path=source, output_path=output, dedupe_rows=True, report_path=report_path))

    payload = json.loads(report_path.read_text(encoding="utf-8"))
    assert payload["summary"]["deleted_duplicate_rows"] == 1
    assert payload["sheets"][0]["sheet"] == "Orders"


def test_duplicate_rows_keep_first_occurrence(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.append(["name", "amount"])
    ws.append(["A", 1])
    ws.append(["A", 1])
    ws.append(["B", 2])

    deleted = delete_duplicate_rows(ws, header_row=1)

    assert deleted == 1
    assert ws.max_row == 3
    assert ws["A2"].value == "A"


def test_is_blank():
    assert is_blank(None)
    assert is_blank("   ")
    assert not is_blank("  x  ")


def test_cli_runs(tmp_path: Path):
    source = tmp_path / "dirty.xlsx"
    output = tmp_path / "clean.xlsx"
    make_dirty_workbook(source)

    assert main([str(source), "-o", str(output), "--dedupe"]) == 0
    assert output.exists()
