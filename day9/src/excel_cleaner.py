#!/usr/bin/env python3
"""Clean common messy Excel workbooks safely."""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


@dataclass
class SheetReport:
    sheet: str
    trimmed_cells: int = 0
    normalized_newlines: int = 0
    deleted_blank_rows: int = 0
    deleted_blank_columns: int = 0
    deleted_duplicate_rows: int = 0


@dataclass
class CleanReport:
    input: str
    output: str
    sheets: list[SheetReport]

    @property
    def total_trimmed_cells(self) -> int:
        return sum(sheet.trimmed_cells for sheet in self.sheets)

    @property
    def total_deleted_blank_rows(self) -> int:
        return sum(sheet.deleted_blank_rows for sheet in self.sheets)

    @property
    def total_deleted_blank_columns(self) -> int:
        return sum(sheet.deleted_blank_columns for sheet in self.sheets)

    @property
    def total_deleted_duplicate_rows(self) -> int:
        return sum(sheet.deleted_duplicate_rows for sheet in self.sheets)


@dataclass(frozen=True)
class CleanOptions:
    input_path: Path
    output_path: Path
    trim_text: bool = True
    normalize_newlines: bool = True
    remove_blank_rows: bool = True
    remove_blank_columns: bool = True
    dedupe_rows: bool = False
    header_row: int = 1
    report_path: Path | None = None
    overwrite: bool = False


def clean_workbook(options: CleanOptions) -> CleanReport:
    if not options.input_path.exists():
        raise FileNotFoundError(f"文件不存在: {options.input_path}")
    if options.output_path.exists() and not options.overwrite:
        raise FileExistsError(f"输出文件已存在，使用 --overwrite 覆盖: {options.output_path}")

    workbook = load_workbook(options.input_path)
    reports: list[SheetReport] = []
    for worksheet in workbook.worksheets:
        reports.append(clean_sheet(worksheet, options))

    options.output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(options.output_path)

    report = CleanReport(
        input=str(options.input_path),
        output=str(options.output_path),
        sheets=reports,
    )
    if options.report_path:
        write_report(report, options.report_path)
    return report


def clean_sheet(worksheet: Worksheet, options: CleanOptions) -> SheetReport:
    report = SheetReport(sheet=worksheet.title)

    if options.trim_text or options.normalize_newlines:
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                cleaned = value
                if options.normalize_newlines:
                    normalized = re.sub(r"\r\n?|\n", " ", cleaned)
                    normalized = re.sub(r"\s+", " ", normalized)
                    if normalized != cleaned:
                        report.normalized_newlines += 1
                    cleaned = normalized
                if options.trim_text:
                    trimmed = cleaned.strip()
                    if trimmed != cleaned:
                        report.trimmed_cells += 1
                    cleaned = trimmed
                if cleaned != value:
                    cell.value = cleaned

    if options.remove_blank_rows:
        report.deleted_blank_rows = delete_blank_rows(worksheet)

    if options.remove_blank_columns:
        report.deleted_blank_columns = delete_blank_columns(worksheet)

    if options.dedupe_rows:
        report.deleted_duplicate_rows = delete_duplicate_rows(worksheet, options.header_row)

    return report


def delete_blank_rows(worksheet: Worksheet) -> int:
    deleted = 0
    for row_index in range(worksheet.max_row, 0, -1):
        if all(is_blank(worksheet.cell(row_index, col).value) for col in range(1, worksheet.max_column + 1)):
            worksheet.delete_rows(row_index)
            deleted += 1
    return deleted


def delete_blank_columns(worksheet: Worksheet) -> int:
    deleted = 0
    for col_index in range(worksheet.max_column, 0, -1):
        if all(is_blank(worksheet.cell(row, col_index).value) for row in range(1, worksheet.max_row + 1)):
            worksheet.delete_cols(col_index)
            deleted += 1
    return deleted


def delete_duplicate_rows(worksheet: Worksheet, header_row: int = 1) -> int:
    seen: set[tuple[Any, ...]] = set()
    deleted = 0
    start_row = max(1, header_row + 1)
    for row_index in range(worksheet.max_row, start_row - 1, -1):
        values = tuple(normalize_for_key(worksheet.cell(row_index, col).value) for col in range(1, worksheet.max_column + 1))
        if all(is_blank(value) for value in values):
            continue
        if values in seen:
            worksheet.delete_rows(row_index)
            deleted += 1
        else:
            seen.add(values)
    return deleted


def normalize_for_key(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def write_report(report: CleanReport, report_path: Path) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    payload = asdict(report)
    payload["summary"] = {
        "trimmed_cells": report.total_trimmed_cells,
        "deleted_blank_rows": report.total_deleted_blank_rows,
        "deleted_blank_columns": report.total_deleted_blank_columns,
        "deleted_duplicate_rows": report.total_deleted_duplicate_rows,
    }
    report_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def print_report(report: CleanReport) -> None:
    print(f"输入: {report.input}")
    print(f"输出: {report.output}")
    for sheet in report.sheets:
        print(f"[{sheet.sheet}]")
        print(f"  修剪文本单元格: {sheet.trimmed_cells}")
        print(f"  规范换行单元格: {sheet.normalized_newlines}")
        print(f"  删除空行: {sheet.deleted_blank_rows}")
        print(f"  删除空列: {sheet.deleted_blank_columns}")
        print(f"  删除重复行: {sheet.deleted_duplicate_rows}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="excel-cleaner",
        description="清理 Excel 表格中的空行空列、文本空格、换行和重复行。",
    )
    parser.add_argument("input", help="输入 .xlsx 文件")
    parser.add_argument("-o", "--output", required=True, help="输出 .xlsx 文件")
    parser.add_argument("--dedupe", action="store_true", help="删除重复数据行，默认保留重复行")
    parser.add_argument("--header-row", type=int, default=1, help="表头所在行，去重时从下一行开始，默认 1")
    parser.add_argument("--keep-blank-rows", action="store_true", help="保留空行")
    parser.add_argument("--keep-blank-cols", action="store_true", help="保留空列")
    parser.add_argument("--no-trim", action="store_true", help="不修剪文本首尾空格")
    parser.add_argument("--keep-newlines", action="store_true", help="保留单元格内换行")
    parser.add_argument("--report", help="输出 JSON 清洗报告")
    parser.add_argument("--overwrite", action="store_true", help="允许覆盖输出文件")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    options = CleanOptions(
        input_path=Path(args.input),
        output_path=Path(args.output),
        trim_text=not args.no_trim,
        normalize_newlines=not args.keep_newlines,
        remove_blank_rows=not args.keep_blank_rows,
        remove_blank_columns=not args.keep_blank_cols,
        dedupe_rows=args.dedupe,
        header_row=args.header_row,
        report_path=Path(args.report) if args.report else None,
        overwrite=args.overwrite,
    )

    try:
        report = clean_workbook(options)
    except (FileNotFoundError, FileExistsError) as exc:
        print(f"错误: {exc}", file=sys.stderr)
        return 1

    print_report(report)
    if options.report_path:
        print(f"报告: {options.report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
